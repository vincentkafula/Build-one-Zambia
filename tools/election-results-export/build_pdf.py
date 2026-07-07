import sys
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                 Spacer, HRFlowable)
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from reportlab.pdfgen import canvas as canvas_mod

# Usage: python build_pdf.py <formatted.xlsx> [output.pdf]
# Expects the workbook produced by build_xlsx.py (Summary + Results sheets
# with the stat/winner cells in the fixed layout it writes).
SRC = sys.argv[1] if len(sys.argv) > 1 else "election-results-formatted.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "election-results-template.pdf"

wb = load_workbook(SRC, data_only=True)
summ = wb["Summary"]
res = wb["Results"]

NAVY = colors.HexColor("#1F3864")
NAVY_LIGHT = colors.HexColor("#2E5395")
GOLD = colors.HexColor("#C9A227")
LIGHT_GRAY = colors.HexColor("#F2F2F2")
TEXT = colors.HexColor("#1F1F1F")
MUTED = colors.HexColor("#595959")
GREEN = colors.HexColor("#1E7145")
GOLD_BG = colors.HexColor("#FFF2CC")

# ---- pull data ----
report_title = summ["A2"].value
generated = summ["A3"].value.replace("Generated: ", "")
winner_name = summ["A6"].value
winner_votes = summ["C6"].value
winner_share = summ["C7"].value

stats = [
    ("Registered Voters", summ["C10"].value, "{:,.0f}"),
    ("Rejected Ballots", summ["C11"].value, "{:,.0f}"),
    ("Valid Votes", summ["C12"].value, "{:,.0f}"),
    ("Votes Cast", summ["C13"].value, "{:,.0f}"),
    ("Voter Turnout", summ["C14"].value, "{:.1%}"),
    ("Candidates Contesting", summ["C15"].value, "{:.0f}"),
]

rows = []
total_votes = None
for row in res.iter_rows(min_row=2, values_only=True):
    rank, cand, party, votes, share = row
    if cand is None and party == "TOTAL":
        total_votes = votes
        continue
    rows.append((rank, cand, party, votes, share))

# ---- header/footer canvas ----
def header_footer(c: canvas_mod.Canvas, doc):
    c.saveState()
    w, h = letter
    # top banner
    c.setFillColor(NAVY)
    c.rect(0, h - 0.95*inch, w, 0.95*inch, fill=1, stroke=0)
    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 20)
    c.drawCentredString(w/2, h - 0.45*inch, "BOZ ELECTION RESULTS")
    c.setFont("Helvetica", 11)
    c.drawCentredString(w/2, h - 0.72*inch, report_title)
    # footer
    c.setFillColor(MUTED)
    c.setFont("Helvetica", 8)
    c.drawString(0.6*inch, 0.4*inch, f"Generated: {generated}")
    c.drawRightString(w - 0.6*inch, 0.4*inch, f"Page {doc.page}")
    c.setStrokeColor(colors.HexColor("#D9D9D9"))
    c.line(0.6*inch, 0.55*inch, w - 0.6*inch, 0.55*inch)
    c.restoreState()

doc = SimpleDocTemplate(OUT, pagesize=letter,
                         topMargin=1.25*inch, bottomMargin=0.75*inch,
                         leftMargin=0.6*inch, rightMargin=0.6*inch)

story = []

label_style = ParagraphStyle("label", fontName="Helvetica-Bold", fontSize=9,
                              textColor=colors.white, alignment=TA_CENTER)
winner_name_style = ParagraphStyle("wname", fontName="Helvetica-Bold", fontSize=17,
                                    textColor=NAVY, alignment=TA_CENTER)
winner_sub_style = ParagraphStyle("wsub", fontName="Helvetica", fontSize=10,
                                   textColor=MUTED, alignment=TA_CENTER)

# Winner banner
story.append(Spacer(1, 4))
winner_header = Table([[Paragraph("PROJECTED WINNER", label_style)]], colWidths=[6.8*inch])
winner_header.setStyle(TableStyle([
    ("BACKGROUND", (0,0), (-1,-1), GOLD),
    ("TOPPADDING", (0,0), (-1,-1), 5),
    ("BOTTOMPADDING", (0,0), (-1,-1), 5),
]))
story.append(winner_header)

winner_body = Table([[
    Paragraph(winner_name, winner_name_style),
    Paragraph(f"{winner_votes:,.0f} votes<br/>{winner_share:.2%} of vote".replace("<br/>", "<br/>"), winner_sub_style)
]], colWidths=[4.0*inch, 2.8*inch])
winner_body.setStyle(TableStyle([
    ("BACKGROUND", (0,0), (-1,-1), GOLD_BG),
    ("BOX", (0,0), (-1,-1), 0.75, GOLD),
    ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
    ("TOPPADDING", (0,0), (-1,-1), 12),
    ("BOTTOMPADDING", (0,0), (-1,-1), 12),
]))
story.append(winner_body)
story.append(Spacer(1, 16))

# Stats header
stats_header = Table([[Paragraph("ELECTION STATISTICS", label_style)]], colWidths=[6.8*inch])
stats_header.setStyle(TableStyle([
    ("BACKGROUND", (0,0), (-1,-1), NAVY),
    ("TOPPADDING", (0,0), (-1,-1), 5),
    ("BOTTOMPADDING", (0,0), (-1,-1), 5),
]))
story.append(stats_header)

stat_label_style = ParagraphStyle("slabel", fontName="Helvetica", fontSize=9.5, textColor=TEXT)
stat_val_style = ParagraphStyle("sval", fontName="Helvetica-Bold", fontSize=10.5, textColor=NAVY, alignment=TA_RIGHT)

stat_rows = []
for i in range(0, len(stats), 2):
    pair = stats[i:i+2]
    row_cells = []
    for label, val, fmt in pair:
        cell_table = Table([[Paragraph(label, stat_label_style), Paragraph(fmt.format(val), stat_val_style)]],
                            colWidths=[2.0*inch, 1.4*inch])
        cell_table.setStyle(TableStyle([
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ]))
        row_cells.append(cell_table)
    stat_rows.append(row_cells)

stats_table = Table(stat_rows, colWidths=[3.4*inch, 3.4*inch])
style_cmds = [
    ("BOX", (0,0), (-1,-1), 0.5, colors.HexColor("#D9D9D9")),
    ("INNERGRID", (0,0), (-1,-1), 0.5, colors.HexColor("#D9D9D9")),
    ("TOPPADDING", (0,0), (-1,-1), 8),
    ("BOTTOMPADDING", (0,0), (-1,-1), 8),
    ("LEFTPADDING", (0,0), (-1,-1), 10),
    ("RIGHTPADDING", (0,0), (-1,-1), 10),
]
for ridx in range(len(stat_rows)):
    bg = LIGHT_GRAY if ridx % 2 == 0 else colors.white
    style_cmds.append(("BACKGROUND", (0, ridx), (-1, ridx), bg))
stats_table.setStyle(TableStyle(style_cmds))
story.append(stats_table)
story.append(Spacer(1, 18))

# Results table header
res_header = Table([[Paragraph("FULL RESULTS", label_style)]], colWidths=[6.8*inch])
res_header.setStyle(TableStyle([
    ("BACKGROUND", (0,0), (-1,-1), NAVY),
    ("TOPPADDING", (0,0), (-1,-1), 5),
    ("BOTTOMPADDING", (0,0), (-1,-1), 5),
]))
story.append(res_header)
story.append(Spacer(1, 6))

col_header_style = ParagraphStyle("colh", fontName="Helvetica-Bold", fontSize=9.5, textColor=colors.white, alignment=TA_CENTER)
cell_style = ParagraphStyle("cell", fontName="Helvetica", fontSize=9.5, textColor=TEXT)
cell_style_r = ParagraphStyle("cellr", fontName="Helvetica", fontSize=9.5, textColor=TEXT, alignment=TA_RIGHT)
cell_style_c = ParagraphStyle("cellc", fontName="Helvetica", fontSize=9.5, textColor=TEXT, alignment=TA_CENTER)
winner_cell_style = ParagraphStyle("wcell", fontName="Helvetica-Bold", fontSize=9.5, textColor=NAVY)
winner_cell_style_r = ParagraphStyle("wcellr", fontName="Helvetica-Bold", fontSize=9.5, textColor=NAVY, alignment=TA_RIGHT)
winner_cell_style_c = ParagraphStyle("wcellc", fontName="Helvetica-Bold", fontSize=9.5, textColor=NAVY, alignment=TA_CENTER)

table_data = [[
    Paragraph("Rank", col_header_style), Paragraph("Candidate", col_header_style),
    Paragraph("Party", col_header_style), Paragraph("Votes", col_header_style),
    Paragraph("Vote Share", col_header_style)
]]
for rank, cand, party, votes, share in rows:
    is_winner = (rank == 1)
    ls = winner_cell_style if is_winner else cell_style
    lc = winner_cell_style_c if is_winner else cell_style_c
    lr = winner_cell_style_r if is_winner else cell_style_r
    table_data.append([
        Paragraph(str(rank), lc),
        Paragraph(str(cand), ls),
        Paragraph(str(party), lc),
        Paragraph(f"{votes:,.0f}", lr),
        Paragraph(f"{share:.2%}", lr),
    ])
table_data.append([
    "", "", Paragraph("TOTAL", ParagraphStyle("tot", fontName="Helvetica-Bold", fontSize=9.5, textColor=colors.white, alignment=TA_RIGHT)),
    Paragraph(f"{total_votes:,.0f}", ParagraphStyle("totv", fontName="Helvetica-Bold", fontSize=9.5, textColor=colors.white, alignment=TA_RIGHT)),
    Paragraph("100.00%", ParagraphStyle("tots", fontName="Helvetica-Bold", fontSize=9.5, textColor=colors.white, alignment=TA_RIGHT)),
])

results_table = Table(table_data, colWidths=[0.6*inch, 2.6*inch, 1.0*inch, 1.2*inch, 1.4*inch], repeatRows=1)
res_style = [
    ("BACKGROUND", (0,0), (-1,0), NAVY),
    ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
    ("TOPPADDING", (0,0), (-1,-1), 5),
    ("BOTTOMPADDING", (0,0), (-1,-1), 5),
    ("LINEBELOW", (0,0), (-1,-2), 0.4, colors.HexColor("#D9D9D9")),
    ("BACKGROUND", (0,-1), (-1,-1), NAVY_LIGHT),
]
for i in range(1, len(table_data)-1):
    if rows[i-1][0] == 1:
        res_style.append(("BACKGROUND", (0,i), (-1,i), GOLD_BG))
    elif i % 2 == 0:
        res_style.append(("BACKGROUND", (0,i), (-1,i), LIGHT_GRAY))
results_table.setStyle(TableStyle(res_style))
story.append(results_table)

doc.build(story, onFirstPage=header_footer, onLaterPages=header_footer)
print("saved", OUT)
