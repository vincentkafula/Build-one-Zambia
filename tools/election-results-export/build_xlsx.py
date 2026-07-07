import sys
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule
from openpyxl.worksheet.pagebreak import Break

# Usage: python build_xlsx.py <input.xlsx> [output.xlsx]
# Input must have a "Summary" sheet (Report Title/Location/Generated/Registered
# Voters/Rejected Ballots key-value rows) and a "Results" sheet
# (Rank, Candidate, Party, Votes, Vote Share (%) columns), matching the raw
# export produced by the results website.
SRC = sys.argv[1] if len(sys.argv) > 1 else "source.xlsx"
OUT = sys.argv[2] if len(sys.argv) > 2 else "election-results-formatted.xlsx"

wb = load_workbook(SRC)
summ = wb["Summary"]
res = wb["Results"]

# ---------- Palette ----------
NAVY = "1F3864"
NAVY_LIGHT = "2E5395"
GOLD = "C9A227"
LIGHT_GRAY = "F2F2F2"
MID_GRAY = "D9D9D9"
WHITE = "FFFFFF"
TEXT = "1F1F1F"
MUTED = "595959"
GREEN = "1E7145"

FONT_NAME = "Arial"

thin = Side(style="thin", color=MID_GRAY)
box_border = Border(left=thin, right=thin, top=thin, bottom=thin)

# capture raw values we need before wiping
report_title = summ["B2"].value
location = summ["B3"].value
generated = summ["B4"].value
registered = summ["B20"].value
rejected = summ["B23"].value

n_candidates = res.max_row - 1  # minus header, includes TOTAL row currently
last_data_row = None
for r in range(2, res.max_row + 1):
    if res.cell(r, 3).value == "TOTAL":
        last_data_row = r - 1
        break
if last_data_row is None:
    last_data_row = res.max_row

# ================= RESULTS SHEET =================
res_max_row_orig = res.max_row
total_row = last_data_row + 1

# widths
res.column_dimensions["A"].width = 8
res.column_dimensions["B"].width = 30
res.column_dimensions["C"].width = 14
res.column_dimensions["D"].width = 14
res.column_dimensions["E"].width = 16

# header row style
header_fill = PatternFill("solid", fgColor=NAVY)
for c in range(1, 6):
    cell = res.cell(1, c)
    cell.font = Font(name=FONT_NAME, size=11, bold=True, color=WHITE)
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = box_border
res.row_dimensions[1].height = 22

# data rows
for r in range(2, last_data_row + 1):
    is_stripe = (r % 2 == 0)
    fill = PatternFill("solid", fgColor=LIGHT_GRAY) if is_stripe else PatternFill("solid", fgColor=WHITE)
    is_winner = (res.cell(r, 1).value == 1)
    for c in range(1, 6):
        cell = res.cell(r, c)
        cell.font = Font(name=FONT_NAME, size=11, bold=is_winner,
                          color=NAVY if is_winner else TEXT)
        cell.fill = PatternFill("solid", fgColor="FFF2CC") if is_winner else fill
        cell.border = box_border
        if c == 1:
            cell.alignment = Alignment(horizontal="center")
        elif c == 4:
            cell.alignment = Alignment(horizontal="right")
            cell.number_format = "#,##0"
        elif c == 5:
            cell.alignment = Alignment(horizontal="right")
            cell.value = f"={cell.coordinate.replace(cell.column_letter,'D')}/D${total_row}"
            cell.number_format = "0.00%"
        else:
            cell.alignment = Alignment(horizontal="left")

# TOTAL row
for c in range(1, 6):
    cell = res.cell(total_row, c)
    cell.font = Font(name=FONT_NAME, size=11, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY_LIGHT)
    cell.border = box_border
res.cell(total_row, 1).value = None
res.cell(total_row, 2).value = None
res.cell(total_row, 3).value = "TOTAL"
res.cell(total_row, 3).alignment = Alignment(horizontal="right")
res.cell(total_row, 4).value = f"=SUM(D2:D{last_data_row})"
res.cell(total_row, 4).number_format = "#,##0"
res.cell(total_row, 4).alignment = Alignment(horizontal="right")
res.cell(total_row, 5).value = f"=D{total_row}/D{total_row}"
res.cell(total_row, 5).number_format = "0.0%"
res.cell(total_row, 5).alignment = Alignment(horizontal="right")

# remove any leftover rows below total (old blank/total row artifacts)
if res.max_row > total_row:
    res.delete_rows(total_row + 1, res.max_row - total_row)

# data bar on vote share
rule = DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1,
                    color="2E5395", showValue=True, minLength=None, maxLength=None)
res.conditional_formatting.add(f"E2:E{last_data_row}", rule)

# freeze header + autofilter
res.freeze_panes = "A2"
res.auto_filter.ref = f"A1:E{total_row}"
res.sheet_view.showGridLines = False

# ================= SUMMARY SHEET =================
# wipe existing content
for row in summ.iter_rows():
    for cell in row:
        cell.value = None
        cell.fill = PatternFill(fill_type=None)
        cell.font = Font(name=FONT_NAME)
        cell.border = Border()
        cell.alignment = Alignment()

summ.column_dimensions["A"].width = 26
summ.column_dimensions["B"].width = 26
summ.column_dimensions["C"].width = 26
summ.column_dimensions["D"].width = 26

# Title banner (rows 1-2 merged across A:D)
summ.merge_cells("A1:D1")
c = summ["A1"]
c.value = "BOZ ELECTION RESULTS"
c.font = Font(name=FONT_NAME, size=20, bold=True, color=WHITE)
c.fill = PatternFill("solid", fgColor=NAVY)
c.alignment = Alignment(horizontal="center", vertical="center")
summ.row_dimensions[1].height = 34

summ.merge_cells("A2:D2")
c = summ["A2"]
c.value = f"{report_title}  |  {location}"
c.font = Font(name=FONT_NAME, size=12, bold=True, color=WHITE)
c.fill = PatternFill("solid", fgColor=NAVY_LIGHT)
c.alignment = Alignment(horizontal="center", vertical="center")
summ.row_dimensions[2].height = 22

summ.merge_cells("A3:D3")
c = summ["A3"]
c.value = f"Generated: {generated}"
c.font = Font(name=FONT_NAME, size=9, italic=True, color=MUTED)
c.alignment = Alignment(horizontal="center")
summ.row_dimensions[3].height = 16

# spacer
summ.row_dimensions[4].height = 8

# Winner banner (rows 5-7)
summ.merge_cells("A5:D5")
c = summ["A5"]
c.value = "PROJECTED WINNER"
c.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
c.fill = PatternFill("solid", fgColor=GOLD)
c.alignment = Alignment(horizontal="center")
summ.row_dimensions[5].height = 18

summ.merge_cells("A6:B7")
c = summ["A6"]
c.value = "=INDEX(Results!B:B,MATCH(1,Results!A:A,0))"
c.font = Font(name=FONT_NAME, size=16, bold=True, color=NAVY)
c.alignment = Alignment(horizontal="center", vertical="center")
c.fill = PatternFill("solid", fgColor="FFF2CC")
for coord in ["A6","B6","A7","B7"]:
    summ[coord].border = box_border
    summ[coord].fill = PatternFill("solid", fgColor="FFF2CC")

summ.merge_cells("C6:D6")
c = summ["C6"]
c.value = "=INDEX(Results!D:D,MATCH(1,Results!A:A,0))"
c.number_format = '#,##0" votes"'
c.font = Font(name=FONT_NAME, size=13, bold=True, color=GREEN)
c.alignment = Alignment(horizontal="center", vertical="center")
c.fill = PatternFill("solid", fgColor="FFF2CC")

summ.merge_cells("C7:D7")
c = summ["C7"]
c.value = "=INDEX(Results!E:E,MATCH(1,Results!A:A,0))"
c.number_format = '0.00%" of vote"'
c.font = Font(name=FONT_NAME, size=11, color=MUTED)
c.alignment = Alignment(horizontal="center", vertical="center")
c.fill = PatternFill("solid", fgColor="FFF2CC")
for coord in ["C6","D6","C7","D7"]:
    summ[coord].border = box_border

summ.row_dimensions[6].height = 26
summ.row_dimensions[7].height = 20

summ.row_dimensions[8].height = 8

# Election stats header
summ.merge_cells("A9:D9")
c = summ["A9"]
c.value = "ELECTION STATISTICS"
c.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
c.fill = PatternFill("solid", fgColor=NAVY)
c.alignment = Alignment(horizontal="center")
summ.row_dimensions[9].height = 18

stats = [
    ("Registered Voters", registered, "#,##0", True),
    ("Rejected Ballots", rejected, "#,##0", True),
    ("Valid Votes", "=Results!$D${}".format(total_row), "#,##0", False),
    ("Votes Cast", "=C12+C11", "#,##0", False),
    ("Voter Turnout", "=C13/C10", "0.0%", False),
    ("Candidates Contesting", f"=COUNT(Results!A2:A{last_data_row})", "0", False),
]
r = 10
for label, val, numfmt, is_input in stats:
    summ.merge_cells(f"A{r}:B{r}")
    lab = summ[f"A{r}"]
    lab.value = label
    lab.font = Font(name=FONT_NAME, size=11, color=TEXT)
    lab.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    lab.fill = PatternFill("solid", fgColor=LIGHT_GRAY if r % 2 == 0 else WHITE)
    lab.border = box_border

    summ.merge_cells(f"C{r}:D{r}")
    val_cell = summ[f"C{r}"]
    val_cell.value = val
    val_cell.number_format = numfmt
    val_cell.font = Font(name=FONT_NAME, size=11, bold=True,
                          color="0000FF" if is_input else NAVY)
    val_cell.alignment = Alignment(horizontal="right", vertical="center", indent=1)
    val_cell.fill = PatternFill("solid", fgColor=LIGHT_GRAY if r % 2 == 0 else WHITE)
    val_cell.border = box_border
    summ.row_dimensions[r].height = 20
    r += 1

summ.row_dimensions[r].height = 8
r += 1

# Top 3 candidates mini-table
summ.merge_cells(f"A{r}:D{r}")
c = summ[f"A{r}"]
c.value = "TOP CANDIDATES"
c.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
c.fill = PatternFill("solid", fgColor=NAVY)
c.alignment = Alignment(horizontal="center")
summ.row_dimensions[r].height = 18
r += 1

hdr_row = r
headers = ["Rank", "Candidate", "Votes", "Share"]
col_map = ["A", "B", "C", "D"]
for i, h in enumerate(headers):
    cell = summ[f"{col_map[i]}{hdr_row}"]
    cell.value = h
    cell.font = Font(name=FONT_NAME, size=10, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=NAVY_LIGHT)
    cell.alignment = Alignment(horizontal="center")
    cell.border = box_border
r += 1

for rank in range(1, 4):
    row = r
    stripe = PatternFill("solid", fgColor=LIGHT_GRAY if rank % 2 == 0 else WHITE)
    a = summ[f"A{row}"]; a.value = rank
    a.alignment = Alignment(horizontal="center"); a.font = Font(name=FONT_NAME, size=10)
    b = summ[f"B{row}"]; b.value = f"=IFERROR(INDEX(Results!B:B,MATCH({rank},Results!A:A,0)),\"\")"
    b.font = Font(name=FONT_NAME, size=10)
    cvotes = summ[f"C{row}"]; cvotes.value = f"=IFERROR(INDEX(Results!D:D,MATCH({rank},Results!A:A,0)),\"\")"
    cvotes.number_format = "#,##0"; cvotes.alignment = Alignment(horizontal="right")
    cvotes.font = Font(name=FONT_NAME, size=10)
    d = summ[f"D{row}"]; d.value = f"=IFERROR(INDEX(Results!E:E,MATCH({rank},Results!A:A,0)),\"\")"
    d.number_format = "0.00%"; d.alignment = Alignment(horizontal="right")
    d.font = Font(name=FONT_NAME, size=10)
    for cell in (a, b, cvotes, d):
        cell.fill = stripe
        cell.border = box_border
    r += 1

summ.sheet_view.showGridLines = False

# ---- Page setup for clean printing/PDF export ----
summ.page_setup.orientation = "portrait"
summ.page_setup.fitToWidth = 1
summ.page_setup.fitToHeight = 1
summ.sheet_properties.pageSetUpPr.fitToPage = True
summ.print_area = f"A1:D{r-1}"
summ.page_margins.left = 0.5
summ.page_margins.right = 0.5
summ.page_margins.top = 0.6
summ.page_margins.bottom = 0.6

res.page_setup.orientation = "portrait"
res.page_setup.fitToWidth = 1
res.page_setup.fitToHeight = 1
res.sheet_properties.pageSetUpPr.fitToPage = True
res.print_area = f"A1:E{total_row}"
res.page_margins.left = 0.5
res.page_margins.right = 0.5
res.page_margins.top = 0.6
res.page_margins.bottom = 0.6

wb.save(OUT)
print("saved", OUT, "total_row", total_row, "last_data_row", last_data_row)
